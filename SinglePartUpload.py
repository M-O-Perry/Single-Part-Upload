from tkinter import filedialog, messagebox
import tkinter as tk
import openpyxl
import os
from EVOUtil import abreviateWords, isNone, createNewPart
from FindVendor import findVendor


file = filedialog.askopenfilename()

if not file:
    print("No file selected")
    os._exit(0)
    
print("File selected: ", file)

ws = openpyxl.load_workbook(file).active

"""
    info that i need
    
    partNumber
    description
    partClass
    partType
    mfg
    mfgNumber
    vendor
    vendorNumber
    specs
        
"""

partNumber = ws["A1"].value
description = ws["B1"].value
partClass = ws["C1"].value
partType = ws["D1"].value
revision = ws["E1"].value
specs = ws["F1"].value
mfg = ws["G1"].value
mfgNumber = ws["H1"].value
vendor = ws["I1"].value
vendorNumber = ws["J1"].value

if not partNumber or not description:
    messagebox.showerror("Error", "Part Number and Description are required")
    os._exit(0)
    

description = abreviateWords(description)
specs = abreviateWords(specs)

if isNone(specs):
    specs = ""
if isNone(vendor):
    vendor = ""
if isNone(vendorNumber):
    vendorPartNumber = ""
if isNone(mfg):
    mfg = ""
if isNone(mfgNumber):
    mfgPartNumber = ""
if isNone(partClass):
    partClass = ""
if isNone(partType):
    partType = ""
if isNone(revision):
    revision = ""


partNumber = partNumber.strip()
description = description.strip()
partClass = partClass.strip()
partType = partType.strip()
mfg = mfg.strip()
mfgNumber = mfgNumber.strip()
vendor = findVendor(vendor.strip()[:4])
vendorNumber = vendorNumber.strip()
revision = revision.strip()



specs = specs.strip()

middle2Digits = partNumber[5:7]
print("middle2Digits", middle2Digits)
print("partNumber", partNumber)

classTypes = {
    "1": "FINM",
    "2": "ASSL",
    "3": "MECH",
    "4": "MSCE",
    "5": "MSCE",
    "6": "MECH",
    "7": "MECH",
    "8": "MSCE",
    "9": "ELEC",
}

if mfg == "" and mfgNumber == "" and vendor == "" and vendorNumber == "":
    partType = "A"
elif mfg != "" and mfgNumber != "" and vendor != "" and vendorNumber != "": 
    partType = "R"
    
else:
    missingInfo = []
    if mfg == "":
        missingInfo.append("Manufacturer")
    if mfgNumber == "":
        missingInfo.append("Manufacturer Number")
    if vendor == "":
        missingInfo.append("Vendor")
    if vendorNumber == "":
        missingInfo.append("Vendor Number")
        
    raise_error_message = f"Error: Incomplete information for part creation, part: {partNumber}.\n Missing information: {", ".join(missingInfo)}"
    
    
    root = tk.Tk()
    root.withdraw()
    tk.messagebox.showerror("Error", raise_error_message)
    
    
    if vendor == "" and vendor != "":
        tk.messagebox.showerror("Error", "Vendor not found in vendor list.")
        
    
    
    root.destroy()
    
    partType = "R"

if partClass == "":
    print(middle2Digits)
    partClass = classTypes[middle2Digits[0]]


createNewPart(partNumber, description, partClass, partType, mfg, mfgNumber, vendor, vendorNumber, specs, revision)

messagebox.showinfo("Success", "Part has been created")